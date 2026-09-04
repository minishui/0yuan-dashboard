import pandas as pd, json, datetime, os, glob, re, sys
import openpyxl

# ===== 路径与冻结配置 =====
BASE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(BASE, 'js', 'data.js')
FROZEN_PATH = os.path.join(BASE, 'frozen_baichuan.json')
# 冻结切割点：< 此日期的数据冻结到 FROZEN_PATH，之后只下载增量底表（9 月起）。
# 与 freeze_data 共用，改这里要同步改 freeze 逻辑里的判断。
FREEZE_CUTOFF = '2026-09-01'
FROZEN_LABEL = '2026-03 ~ 2026-08'

# ===== 动态识别最新百川底表 =====
# 扫描 ~/Downloads 下「百川数据看板*2026-*.xlsx」，按文件名日期（YYYY-MM-DD）取最新，
# 同日多个文件则取修改时间(mtime)最新。兜底：无日期匹配时取 mtime 最新。
def _has_main_funnel_cols(f):
    """落地页主链路底表必须含 pagekey 列（点击趋势等异类报表无此列）"""
    try:
        wb = openpyxl.load_workbook(f, read_only=True)
        ws = wb[wb.sheetnames[0]]
        cols = [c.value for c in next(ws.iter_rows(max_row=1))]
        wb.close()
        return 'pagekey' in cols
    except Exception:
        return False

def find_latest_baichuan():
    pattern = os.path.expanduser('~/Downloads/百川数据看板*2026-*.xlsx')
    files = glob.glob(pattern)
    if not files:
        # 兜底：任意百川 xlsx
        files = glob.glob(os.path.expanduser('~/Downloads/百川数据看板*.xlsx'))
    if not files:
        raise SystemExit('未找到百川底表（~/Downloads/百川数据看板*2026-*.xlsx）')
    # 优先筛选含 pagekey 主链路结构的底表，避免被「点击趋势」等异类报表（按日期更新）鸠占鹊巢
    valid = [f for f in files if _has_main_funnel_cols(f)]
    if valid:
        files = valid
    date_re = re.compile(r'(\d{4})-(\d{2})-(\d{2})')
    def key(f):
        m = date_re.search(os.path.basename(f))
        d = m.group(0) if m else '0000-00-00'
        return (d, os.path.getmtime(f))
    files.sort(key=key)
    return files[-1]

SRC = find_latest_baichuan()

# 新底表(2026-08-20)事件名称为中文显示名，且缺少旧的「端」「用户类型」两列。
# 这里把中文名映射回看板 config.EVENT_MAP 依赖的标准英文事件码；
# 未列出的事件名（如实时加微曝光、switchPhone* 等）保留原名，看板不引用、无害。
def map_code(x):
    x = str(x).strip()
    exact = {
        '展现PV': 'firstPagePv',
        'slideScreen': 'slideScreen',
        '选择年级': 'choosegrade',
        '授权登录成功': 'authorizeLoginSuccess',
        '选择课程': 'choosecourse',
        'zero_order_success': 'zero_order_success',
        '获取课程详情后，无可选课程': 'notFoundOptionalCourse',
        '下单失败': 'createordererror',
        '用户拒绝获取手机号': '38',
        '用户授权同意获取手机号': '39',
        '授权注册成功': '90',
        '手机号授权登录失败': 'authorizeLoginFail',
        '用户是否登录': 'loginstate',
        'getTokenWithCode': 'getTokenWithCode',
        '选择课程弹窗显示了': 'coursePopIsShown',
        '选择年级半弹窗展示': 'gradehalftoast',
        '点击立即购买': 'buy',  # 原双入口事件，保留数据不进漏斗
    }
    if x in exact:
        return exact[x]
    if '立即支付' in x:          # 提交支付信息【点击“立即支付”时】
        return 'paysubmit'
    return x

READ_KEYWORDS = ['阅读', '经典', '茅盾', '诵读', '征文', '全球青少年']
def biz_line(name):
    for k in READ_KEYWORDS:
        if k in name:
            return '阅读'
    return '数学'

def read_baichuan_df(src):
    """读取底表并做标准化（中文事件名 -> 标准事件码、剔除空事件名、日期格式化）"""
    df = pd.read_excel(src, sheet_name='sheet',
        usecols=['事件日期', '页面名称', '事件名称', 'pagekey', 'uv'],
        engine='openpyxl')
    df = df[df['事件名称'].notna()].copy()
    df.rename(columns={'事件名称': '事件标识'}, inplace=True)
    df['事件标识'] = [map_code(x) for x in df['事件标识']]
    df['uv'] = df['uv'].fillna(0).astype(int)
    df['事件日期'] = pd.to_datetime(df['事件日期']).dt.strftime('%Y-%m-%d')
    return df

def aggregate(df):
    """把底表子集聚合为 facts_list / pagekeys / dates / all_codes（与 0 元 frozen 同构）"""
    pk_name = df.groupby(['pagekey', '页面名称']).size().reset_index(name='c')
    pk_name = pk_name.sort_values('c', ascending=False).drop_duplicates('pagekey')
    pk2name = dict(zip(pk_name['pagekey'], pk_name['页面名称']))
    pk2bl = {pk: biz_line(nm) for pk, nm in pk2name.items()}

    facts = {}
    for (pk, d), g in df.groupby(['pagekey', '事件日期']):
        ev = {}
        for code, uv in zip(g['事件标识'], g['uv']):
            uv = int(uv)
            if uv > 0:
                ev[code] = ev.get(code, 0) + uv
        if not ev:
            continue
        facts[(pk, d)] = ev
    facts_list = [{'pk': pk, 'd': d, 'e': '全部', 'ut': '全部', 'ev': ev}
                  for (pk, d), ev in facts.items()]

    visit_by_pk = df[df['事件标识'] == 'firstPagePv'].groupby('pagekey')['uv'].sum().to_dict()
    pagekeys = [{'pk': pk, 'nm': pk2name.get(pk, pk), 'bl': pk2bl.get(pk, '数学'),
                 'visit_uv': int(visit_by_pk.get(pk, 0))}
                for pk in sorted(df['pagekey'].unique())]
    pagekeys.sort(key=lambda x: -x['visit_uv'])

    dates = sorted(df['事件日期'].unique())
    all_codes = sorted(set(df['事件标识'].tolist()))
    return facts_list, pagekeys, dates, all_codes

def load_frozen():
    if not os.path.exists(FROZEN_PATH):
        return None
    try:
        with open(FROZEN_PATH, encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return None

# ===== 模式 1：冻结（仅首次/数据校正时运行）=====
def cmd_freeze():
    print(f"[freeze] 读取底表并冻结 < {FREEZE_CUTOFF} 的数据 ...")
    df = read_baichuan_df(SRC)
    fr = df[df['事件日期'] < FREEZE_CUTOFF]
    if fr.empty:
        raise SystemExit(f'[freeze] 底表无 < {FREEZE_CUTOFF} 的数据，无需冻结')
    facts_list, pagekeys, dates, all_codes = aggregate(fr)
    frozen = {
        'frozen_date_min': dates[0],
        'frozen_date_max': dates[-1],
        'cutoff': FREEZE_CUTOFF,
        'source': os.path.basename(SRC),
        'generated_at': datetime.datetime.now().strftime('%Y-%m-%d %H:%M'),
        'pagekeys': pagekeys,
        'facts': facts_list,
    }
    with open(FROZEN_PATH, 'w', encoding='utf-8') as f:
        json.dump(frozen, f, ensure_ascii=False, separators=(',', ':'))
    print(f"[freeze] 冻结区间 {dates[0]} ~ {dates[-1]} | facts {len(facts_list):,} | pagekey {len(pagekeys)} -> {FROZEN_PATH}")

# ===== 模式 2：生成 data.js（每次自动化调用）=====
def cmd_build():
    frozen = load_frozen()
    if frozen:
        # 增量底表只取 >= 切割点的数据；3-8 月由冻结文件提供，避免过渡期重复下载导致翻倍
        df = read_baichuan_df(SRC)
        live = df[df['事件日期'] >= FREEZE_CUTOFF]
        live_facts, live_pk, live_dates, live_codes = aggregate(live)

        # 合并：frozen 提供 3-8 月，live 提供 9 月起；按 (pk,d) 自然补全，pagekey 合并求和 visit_uv
        facts = frozen['facts'] + live_facts
        pkmap = {p['pk']: dict(p) for p in frozen['pagekeys']}
        for p in live_pk:
            if p['pk'] in pkmap:
                pkmap[p['pk']]['visit_uv'] += p['visit_uv']
            else:
                pkmap[p['pk']] = dict(p)
        pagekeys = sorted(pkmap.values(), key=lambda x: -x['visit_uv'])

        fz_dates = [f['d'] for f in frozen['facts']]
        dates = sorted(set(fz_dates) | set(live_dates))
        all_codes = sorted(set(live_codes) | set().union(*[set(f['ev'].keys()) for f in frozen['facts']])) if (live_codes or frozen['facts']) else []

        fmin = frozen.get('frozen_date_min'); fmax = frozen.get('frozen_date_max')
        dmin = min([x for x in [fmin] + dates if x])
        dmax = max([x for x in [fmax] + dates if x])
        frozen_note = f"frozen={FROZEN_LABEL}({len(frozen['facts']):,})"
    else:
        # 未冻结：沿用旧逻辑读全量（不丢数据，仅打印提示）
        print('[warn] 未找到 frozen_baichuan.json，使用全量底表（建议先 --freeze 一次）')
        df = read_baichuan_df(SRC)
        facts, pagekeys, dates, all_codes = aggregate(df)
        dmin, dmax = dates[0], dates[-1]
        frozen_note = 'frozen=未启用'

    meta = {
        'generated_at': datetime.datetime.now().strftime('%Y-%m-%d %H:%M'),
        'source': os.path.basename(SRC),
        'date_min': dmin, 'date_max': dmax,
        'business_lines': ['阅读', '数学'],
        'ends': ['全部'],
        'user_types': ['全部'],
        'pagekeys': pagekeys, 'all_event_codes': all_codes,
        'frozen_source': (frozen['source'] if frozen else None),
        'frozen_range': (f'{frozen["frozen_date_min"]}~{frozen["frozen_date_max"]}' if frozen else None),
    }
    data = {'meta': meta, 'facts': facts}
    with open(OUT, 'w', encoding='utf-8') as f:
        f.write('window.DASHBOARD_DATA = ')
        json.dump(data, f, ensure_ascii=False, separators=(',', ':'))
        f.write(';\n')

    size = os.path.getsize(OUT)
    n_read = sum(1 for p in pagekeys if p['bl'] == '阅读')
    n_math = sum(1 for p in pagekeys if p['bl'] == '数学')
    print(f"facts 条数: {len(facts):,}")
    print(f"pagekey 数: {len(pagekeys)}  (阅读={n_read}, 数学={n_math})  [{frozen_note}]")
    print(f"data.js 大小: {size / 1024 / 1024:.2f} MB -> {OUT}")

if __name__ == '__main__':
    if '--freeze' in sys.argv:
        cmd_freeze()
    else:
        cmd_build()
